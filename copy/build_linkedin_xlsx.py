#!/usr/bin/env python3
"""
Mendel — LinkedIn Copy export to .xlsx
Two sheets:
 - "Por variante": one row per persona variant, all 3 messages + 4 personalization examples
 - "Por mensaje": one row per message (M1, M2, M3) ready for HeyReach upload
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path("/Users/matteofois/Documents/Claude/Claude Folder/clients/mendel/copy/Mendel_LinkedIn_Copy.xlsx")

# Brand colors
NAVY = "1A1A2E"
BLUE = "0F3460"
LGREY = "F3F4F6"
AMBER_FILL = "FFFBEB"
GREEN_FILL = "F0FDF4"
BLUE_FILL = "EFF6FF"
WHITE = "FFFFFF"

# Sequences data — same content as the PDF
SEQUENCES = [
    {
        "angle_num": "1",
        "angle_name": "Recuperación de CFDIs y deducibilidad",
        "variant": "1A",
        "persona": "Gerente de Impuestos / Responsable de Contabilidad",
        "m1": (
            "{{firstName}}, ayudamos a equipos de impuestos en México a recuperar CFDIs "
            "automáticamente y validarlos con el SAT desde el momento del cargo. Solo trabajamos "
            "con operaciones en LatAm. Quería conectar."
        ),
        "m2": (
            "Hola {{firstName}}, gracias por conectar.\n\n"
            "[PERSONALIZACIÓN]\n\n"
            "Lo que escucho seguido de equipos de impuestos en México: el área termina persiguiendo "
            "facturas y armando el reporte de transparencia a mano cada cierre. La deducibilidad que "
            "se pierde no regresa.\n\n"
            "Mendel automatiza la recuperación de CFDIs y la validación con el SAT desde el cargo. "
            "Mercado Libre subió 30% la recuperación de facturas deducibles con esta arquitectura.\n\n"
            "¿Tendría sentido revisarlo para {{companyName}}?"
        ),
        "m3": (
            "Hola {{firstName}}, armé un comparativo de cómo equipos de impuestos en operaciones tipo "
            "{{companyName}} están recuperando 30% más deducibilidad sin sumar personal.\n\n"
            "Tres páginas, datos reales y anónimos. Sin venta.\n\n"
            "¿Te lo paso?\n\n"
            "PD: Mendel se conecta nativo a SAP, Contpaqi y Oracle, con SAT integrado al núcleo."
        ),
        "ex1": "Vi tu publicación sobre [tema fiscal / SAT / CFDI] — coincide con lo que escucho de equipos similares.",
        "ex2": "Noté que {{companyName}} tiene operación multi-entidad — el flujo de CFDIs entre subsidiarias suele ser donde más deducibilidad se pierde.",
        "ex3": "Vi que {{companyName}} expandió a [nuevo mercado / país] — esos movimientos suelen complicar el cierre fiscal.",
        "ex4": "Leí tu comentario sobre [auditoría / cierre / SAT] — es el dolor más común que escucho del área.",
    },
    {
        "angle_num": "1",
        "angle_name": "Recuperación de CFDIs y deducibilidad",
        "variant": "1B",
        "persona": "CFO / Director de Finanzas",
        "m1": (
            "{{firstName}}, ayudamos a CFOs en LatAm a recuperar deducibilidad fiscal y controlar "
            "el gasto antes del cargo. Mercado Libre subió 30% la recuperación de facturas con esta "
            "arquitectura. Quería conectar."
        ),
        "m2": (
            "Hola {{firstName}}, gracias por conectar.\n\n"
            "[PERSONALIZACIÓN]\n\n"
            "El patrón que veo en CFOs en México: entre 20 y 30% de la deducibilidad fiscal se pierde "
            "por CFDIs que nunca llegan o llegan tarde. Eso pega directo al resultado y no se recupera "
            "al cierre.\n\n"
            "Mendel automatiza la recuperación y validación con el SAT desde el cargo. Mercado Libre "
            "subió 30% la recuperación después de reemplazar SAP Concur.\n\n"
            "¿Te resuena para {{companyName}}?"
        ),
        "m3": (
            "Hola {{firstName}}, armé un comparativo de cómo CFOs en operaciones tipo {{companyName}} "
            "están recuperando deducibilidad y cerrando el mes el mismo día.\n\n"
            "Tres páginas, datos reales y anónimos. Sin venta.\n\n"
            "¿Te lo paso?\n\n"
            "PD: Operaciones como Mercado Libre, AB InBev y KPMG corren sobre esta arquitectura."
        ),
        "ex1": "Vi que {{companyName}} acaba de [cerrar ronda / hito / contrato grande] — en ese momento el control del gasto suele entrar al radar del consejo.",
        "ex2": "Noté que entraste recientemente como CFO a {{companyName}} — los primeros 90 días suelen revelar dónde se pierde deducibilidad.",
        "ex3": "Vi tu publicación sobre [finanzas / cierre / flujo de caja] — el punto sobre [X] coincide con lo que escucho de CFOs en operaciones similares.",
        "ex4": "Leí que {{companyName}} expandió a [nuevo mercado / país] — los flujos fiscales multi-país son donde más deducibilidad se filtra.",
    },
    {
        "angle_num": "2",
        "angle_name": "Control en tiempo real (no reportes después del cierre)",
        "variant": "2A",
        "persona": "CFO (frame arquitectónico)",
        "m1": (
            "{{firstName}}, ayudamos a CFOs en LatAm a mover el control del gasto antes del cargo, "
            "no después como Concur. Mercado Libre reemplazó SAP Concur con esta arquitectura. "
            "Quería conectar."
        ),
        "m2": (
            "Hola {{firstName}}, gracias por conectar.\n\n"
            "[PERSONALIZACIÓN]\n\n"
            "Lo que cuentan los CFOs: Concur y AMEX se diseñaron para reportar el gasto después de que "
            "pasa. Sirve si lo que necesitas es visibilidad después del cierre, pero no controla nada "
            "en el momento del cargo.\n\n"
            "Mendel evalúa cada cargo contra política antes de que pase: por empleado, proveedor, "
            "monto y horario. Si viola política, no pasa.\n\n"
            "¿Tendría sentido revisarlo para {{companyName}}?"
        ),
        "m3": (
            "Hola {{firstName}}, armé un comparativo de cómo CFOs en operaciones tipo {{companyName}} "
            "mueven el control antes del cargo, no después.\n\n"
            "Tres páginas, datos reales y anónimos. Sin venta.\n\n"
            "¿Te lo paso?\n\n"
            "PD: Mercado Libre reemplazó SAP Concur con esta arquitectura."
        ),
        "ex1": "Vi tu publicación sobre [cierre mensual / control / gasto] — el punto sobre [X] es exactamente lo que escucho de CFOs en operaciones similares.",
        "ex2": "Noté que {{companyName}} está creciendo rápido el equipo operativo — el desfase entre velocidad de gasto y velocidad de control suele aparecer ahí.",
        "ex3": "Vi que {{companyName}} acaba de [cerrar contrato / firmar acuerdo] — los movimientos grandes revelan huecos de control que no estaban en el radar.",
        "ex4": "Leí que migraron de [herramienta A] a [herramienta B] — me da curiosidad si el control en tiempo real estaba en el alcance del cambio.",
    },
    {
        "angle_num": "2",
        "angle_name": "Control en tiempo real (no reportes después del cierre)",
        "variant": "2B",
        "persona": "Contralor / Director de Finanzas (frame operativo)",
        "m1": (
            "{{firstName}}, ayudamos a Contralores en México a cerrar el mes el mismo día, sin "
            "perseguir tickets ni armar hojas paralelas. Walmart y OXXO corren así. Quería conectar."
        ),
        "m2": (
            "Hola {{firstName}}, gracias por conectar.\n\n"
            "[PERSONALIZACIÓN]\n\n"
            "El patrón que escucho de Contralores en México: el equipo termina persiguiendo tickets y "
            "armando el cierre en Excel mientras Concur reporta lo que ya pasó. Eso suma días al "
            "cierre mensual.\n\n"
            "Mendel aplica política en el momento del cargo, recupera CFDIs automáticamente y "
            "reconcilia con el ERP en tiempo real. El equipo cierra el mes el mismo día.\n\n"
            "¿Hace sentido revisarlo para {{companyName}}?"
        ),
        "m3": (
            "Hola {{firstName}}, armé un comparativo de cómo Contralores en operaciones tipo "
            "{{companyName}} están cerrando el mes el mismo día.\n\n"
            "Tres páginas, datos reales y anónimos. Sin venta.\n\n"
            "¿Te lo paso?\n\n"
            "PD: Walmart y OXXO corren sobre esta arquitectura para sus cadenas multi-tienda."
        ),
        "ex1": "Vi tu publicación sobre [cierre / conciliación / ERP] — el dolor que describes coincide con lo que escucho de Contralores en operaciones similares.",
        "ex2": "Noté que {{companyName}} es operación multi-planta — el cierre multi-entidad suele ser donde más tiempo manual se pierde.",
        "ex3": "Vi que {{companyName}} migró a SAP S/4HANA — el control del gasto conectado nativo al ERP suele ser un hueco visible en esos rollouts.",
        "ex4": "Leí tu comentario sobre [equipo / crecimiento / contratación] — los Contralores que conozco sienten el desfase entre velocidad operativa y velocidad de cierre.",
    },
    {
        "angle_num": "3",
        "angle_name": "Una plataforma, no cuatro herramientas separadas",
        "variant": "3A",
        "persona": "CFO (consolidación)",
        "m1": (
            "{{firstName}}, ayudamos a CFOs en LatAm a consolidar tarjetas, gastos, viajes y ERP en "
            "una sola plataforma, no en cuatro herramientas separadas. KPMG y Mercado Libre corren "
            "así. Quería conectar."
        ),
        "m2": (
            "Hola {{firstName}}, gracias por conectar.\n\n"
            "[PERSONALIZACIÓN]\n\n"
            "El sistema típico que veo: Concur para reportes, tarjetas del banco para gasto, agencia "
            "de viajes para reservas, SAP para ERP. Cada integración entre ellos es un punto de "
            "falla, y cada cierre depende de que los cuatro se conecten sin errores.\n\n"
            "Mendel unifica tarjetas, gastos, viajes, CFDI y ERP en una sola plataforma. KPMG y "
            "Mercado Libre corren así hoy.\n\n"
            "¿Te resuena para {{companyName}}?"
        ),
        "m3": (
            "Hola {{firstName}}, armé un comparativo de cómo CFOs en operaciones tipo {{companyName}} "
            "están consolidando tarjetas, gastos, viajes y ERP en una sola plataforma.\n\n"
            "Tres páginas, datos reales y anónimos. Sin venta.\n\n"
            "¿Te lo paso?\n\n"
            "PD: Operaciones como KPMG, AB InBev y Mercado Libre corren así hoy."
        ),
        "ex1": "Vi que {{companyName}} expandió a [nuevo mercado / país] — los flujos multi-país suelen revelar dónde el sistema actual está atado con cinta.",
        "ex2": "Noté que {{companyName}} usa SAP — la conexión nativa entre gasto y ERP suele ser el punto donde más se rompe la reconciliación.",
        "ex3": "Vi tu publicación sobre [operación / cierre / crecimiento] — el punto sobre [X] coincide con lo que escucho de CFOs consolidando sistemas.",
        "ex4": "Leí que {{companyName}} cerró [contrato grande / acuerdo] — esos eventos exponen huecos entre sistemas que no se veían a menor escala.",
    },
    {
        "angle_num": "3",
        "angle_name": "Una plataforma, no cuatro herramientas separadas",
        "variant": "3B",
        "persona": "Responsable de Viajes (Mendel Viajes)",
        "m1": (
            "{{firstName}}, ayudamos a Responsables de Viajes en LatAm a consolidar reservas, "
            "política y gasto de viajes en una sola plataforma, no entre agencia, correo y "
            "reembolsos. Quería conectar."
        ),
        "m2": (
            "Hola {{firstName}}, gracias por conectar.\n\n"
            "[PERSONALIZACIÓN]\n\n"
            "Lo que escucho de Responsables de Viajes: el sistema típico — agencia de viajes, correo, "
            "hojas de aprobación, reembolsos — se diseñó para reservar y reportar, no para aplicar "
            "política al momento de la reserva. Los viajeros reservan fuera de política y los "
            "reembolsos se aprueban a mano.\n\n"
            "Mendel Viajes integra reserva, política, aprobaciones y gasto en un solo flujo. Los "
            "viajeros reservan dentro de política sin perseguir aprobaciones.\n\n"
            "¿Tiene sentido revisarlo para {{companyName}}?"
        ),
        "m3": (
            "Hola {{firstName}}, armé un comparativo de cómo Responsables de Viajes en operaciones "
            "tipo {{companyName}} están consolidando reservas, política y gasto en una sola "
            "plataforma.\n\n"
            "Tres páginas, datos reales y anónimos. Sin venta.\n\n"
            "¿Te lo paso?\n\n"
            "PD: Firmas multi-país como KPMG corren sobre esta arquitectura."
        ),
        "ex1": "Vi tu publicación sobre [viajes / reservas / política] — el punto sobre [X] coincide con lo que escucho de Responsables de Viajes en operaciones similares.",
        "ex2": "Noté que {{companyName}} es operación multi-país — los flujos de aprobación multi-país son donde más se van los reembolsos manuales.",
        "ex3": "Vi que {{companyName}} acaba de [expandirse / contratar / mover oficinas] — esos eventos suelen disparar volumen de viajes que el sistema actual no maneja bien.",
        "ex4": "Leí que migraron a [agencia X] — me da curiosidad si la política aplicada a la reserva estaba en el alcance del cambio.",
    },
]


def style_header(cell):
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )


def style_body(cell, fill=None):
    cell.font = Font(name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    if fill:
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")


def build_sheet_por_variante(ws):
    headers = [
        "Ángulo", "Variante", "Persona objetivo",
        "Mensaje 1 — Nota de conexión (con la solicitud)",
        "Mensaje 2 — Seguimiento (Día 1-2)",
        "Mensaje 3 — Material de valor (Día 5-7)",
        "Ejemplo personalización 1",
        "Ejemplo personalización 2",
        "Ejemplo personalización 3",
        "Ejemplo personalización 4",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        style_header(c)

    for row_idx, seq in enumerate(SEQUENCES, start=2):
        vals = [
            f"{seq['angle_num']} — {seq['angle_name']}",
            seq["variant"],
            seq["persona"],
            seq["m1"],
            seq["m2"],
            seq["m3"],
            seq["ex1"],
            seq["ex2"],
            seq["ex3"],
            seq["ex4"],
        ]
        for col_idx, v in enumerate(vals, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=v)
            fill = LGREY if row_idx % 2 == 0 else None
            if col_idx == 4:  # Mensaje 1
                fill = BLUE_FILL
            elif col_idx == 5:  # Mensaje 2
                fill = LGREY
            elif col_idx == 6:  # Mensaje 3
                fill = GREEN_FILL
            elif col_idx >= 7:  # Examples
                fill = AMBER_FILL
            style_body(c, fill=fill)

    # Column widths
    col_widths = [22, 10, 32, 50, 70, 60, 50, 50, 50, 50]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row heights — make them generous so the body wraps comfortably
    ws.row_dimensions[1].height = 32
    for r in range(2, len(SEQUENCES) + 2):
        ws.row_dimensions[r].height = 280

    ws.freeze_panes = "D2"


def build_sheet_por_mensaje(ws):
    headers = [
        "Ángulo", "Variante", "Persona objetivo",
        "Tipo de mensaje", "Cuándo enviar", "Texto del mensaje",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        style_header(c)

    row = 2
    for seq in SEQUENCES:
        for tipo, cuando, texto, fill in [
            ("M1 — Nota de conexión", "Con la solicitud",     seq["m1"], BLUE_FILL),
            ("M2 — Seguimiento",      "Día 1-2 tras aceptar", seq["m2"], LGREY),
            ("M3 — Material de valor","Día 5-7",              seq["m3"], GREEN_FILL),
        ]:
            vals = [
                f"{seq['angle_num']} — {seq['angle_name']}",
                seq["variant"],
                seq["persona"],
                tipo,
                cuando,
                texto,
            ]
            for col_idx, v in enumerate(vals, start=1):
                c = ws.cell(row=row, column=col_idx, value=v)
                style_body(c, fill=fill)
            ws.row_dimensions[row].height = 160 if "M2" in tipo else 110
            row += 1

    col_widths = [22, 10, 32, 24, 22, 90]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "D2"


def build_sheet_ejemplos(ws):
    headers = ["Variante", "Persona", "Ejemplo de personalización"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        style_header(c)

    row = 2
    for seq in SEQUENCES:
        for ex_key in ("ex1", "ex2", "ex3", "ex4"):
            vals = [seq["variant"], seq["persona"], seq[ex_key]]
            for col_idx, v in enumerate(vals, start=1):
                c = ws.cell(row=row, column=col_idx, value=v)
                style_body(c, fill=AMBER_FILL)
            ws.row_dimensions[row].height = 50
            row += 1

    col_widths = [12, 32, 110]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def build():
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Por variante"
    build_sheet_por_variante(ws1)

    ws2 = wb.create_sheet(title="Por mensaje")
    build_sheet_por_mensaje(ws2)

    ws3 = wb.create_sheet(title="Ejemplos personalización")
    build_sheet_ejemplos(ws3)

    wb.save(OUTPUT)
    print(f"Built: {OUTPUT}")


if __name__ == "__main__":
    build()
